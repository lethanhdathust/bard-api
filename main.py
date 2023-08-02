# 
from bardapi import Bard,ChatBard
import openpyxl
import streamlit as st
from streamlit_chat import message
import os
import sys
def run_app(message):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Message"
    ws["A2"] = message
    wb.save("answer.xlsx")
    os.system("input.xlsx")

def export_to_excel(answer):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Question", "Answer"])
    ws.append([message, answer])
    wb.save(os.path.join("answer.xlsx"))
    if message == "stop":
        sys.exit()

    
os.environ['_BARD_API_KEY']='ZQiIXFOvIxLqrSSK1EWzmFm7Ee6bgX21Vr2n4aJvenOENuAHvMQGoVJd20CLnQ3L9LC2lA.'
# run_app(message)


st.title("Google bard ")

def response_api(message):
    answer = Bard().get_answer(
 'echo '+f"""Definition of "professional tone": "A professional tone is a way of writing that conveys a sense of formality, respect, and competence. A person writing with a professional tone uses language and intonation that is more formal and appropriate for a business or formal setting. A professional tone can be identified by a number of verbal and nonverbal cues, including:
Use of formal language and vocabulary
Avoidance of slang and colloquial expressions
Appropriate use of titles and honorifics
Direct and concise statements
Maintaining a neutral tone
Use of polite language and manners
Overall, a professional tone communicates a sense of competence and credibility, which can help establish trust and influence in business or formal settings. It is important to note that a professional tone should be tailored to the specific situation and audience, as different contexts may require different levels of formality or informality."
I will give you text content, you will rewrite it and output that in a "professional tone".
Keep the meaning the same. Make sure the re-written content's number of characters is exactly the same as the original text's number of characters. Do not alter the original structure and formatting outlined in any way. Only give me the output and nothing else.
Now, using the concepts above, re-write the following text. Write to make it easy for the audience to read. This is the text i give you to work ‘’{message}‘’  :Must bold more keywords to help the audience.Always do this after rewriting""" 
    )['content']
    return answer
def user_input():
    input_text= st.text_area("Enter your post : ")
    return input_text
if 'generated' not in st.session_state:
    st.session_state['generated']=[]
if 'past' not in st.session_state:
    st.session_state['past']=[]
user_text = user_input()
if user_text:
    output = response_api(user_text)
    st.session_state['generated'].append(output)
    st.session_state['past'].append(user_text)
if st.session_state['generated']:
    for i in range(len (st.session_state['generated']) -1, -1, -1):
        message(st.session_state["past"][i], is_user=True, key=str(i)+'_user')
        message (st.session_state["generated"][i], key=str(i))






# export_to_excel(answer)
