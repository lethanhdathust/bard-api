from bardapi import Bard,ChatBard
import openpyxl
import os
import sys
os.environ['_BARD_API_KEY']='ZQiIXFOvIxLqrSSK1EWzmFm7Ee6bgX21Vr2n4aJvenOENuAHvMQGoVJd20CLnQ3L9LC2lA.'

# def run_app(message):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws["A1"] = "Message"
#     ws["A2"] = message
#     wb.save("okay.xlsx")

    # if message == "stop":
    #     sys.exit()
message = input("Enter your message: ")
answer = Bard().get_answer(
   'echo '+f"""Definition of "professional tone": "A professional tone is a way of writing that conveys a sense of formality, respect, and competence. A person writing with a professional tone uses language and intonation that is more formal and appropriate for a business or formal setting. A professional tone can be identified by a number of verbal and nonverbal cues, including:
Use of formal language and vocabulary
Avoidance of slang and colloquial expressions
Appropriate use of titles and honorifics
Direct and concise statements
Maintaining a neutral tone
Use of polite language and manners
Overall, a professional tone communicates a sense of competence and credibility, which can help establish trust and influence in business or formal settings. It is important to note that a professional tone should be tailored to the specific situation and audience, as different contexts may require different levels of formality or informality."
I will give you text content, you will rewrite it and output that in a "professional tone".
Keep the meaning the same. Make sure the re-written content's number of characters is exactly the same as the original text's number of characters. Do not alter the original structure and formatting outlined in any way. Only give me the output and nothing else.
Now, using the concepts above, re-write the following text. Write to make it easy for the audience to read. This is the text i give you to work ‘’{message}‘’  :Must bold more keywords to help the audience.Always do this after rewriting""" 
    )['content']
print(answer)
# if __name__ == "__main__":
#     main()